from django.contrib import admin
from django import forms
from django.urls import path
from django.http import HttpResponse
from django.apps import apps

import openpyxl
from openpyxl.utils import get_column_letter

from .models import (
    SupportType, Technology, Round, Task,
    Expert, Candidate, Case, PH, Notification, MarketingTeam,Company
)

# ---------- Core model admins ----------

@admin.register(SupportType)
class SupportTypeAdmin(admin.ModelAdmin):
    list_display = ['id', 'support_name']
    search_fields = ['support_name']


@admin.register(Technology)
class TechnologyAdmin(admin.ModelAdmin):
    list_display = ['id', 'technology_name']
    search_fields = ['technology_name']

@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    list_display = ['id', 'company_name']
    search_fields = ['company_name']
    
@admin.register(Round)
class RoundAdmin(admin.ModelAdmin):
    list_display = ['id', 'round_name']
    search_fields = ['round_name']


@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    list_display = ['id', 'task_status']
    list_filter = ['task_status']
    search_fields = ['task_status']


class CandidateInline(admin.TabularInline):
    model = Candidate
    extra = 0
    fields = ['name', 'technology', 'status_flag']
    show_change_link = True


class ExpertAdminForm(forms.ModelForm):
    class Meta:
        model = Expert
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Only experts who are team leads can be selected as team_lead
        self.fields['team_lead'].queryset = Expert.objects.filter(is_team_lead=True)


@admin.register(Expert)
class ExpertAdmin(admin.ModelAdmin):
    form = ExpertAdminForm
    list_display = ['id', 'name', 'is_team_lead', 'team_lead', 'status_flag']
    list_filter = ['is_team_lead', 'status_flag']
    search_fields = ['name']
    inlines = [CandidateInline]


# admin.py
from django import forms
from django.contrib import admin
from .models import Candidate, Case


# ---------- Candidate Admin ----------

@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    list_display = ['id', 'name', 'technology', 'expert_status', 'status_flag']
    list_filter = ['technology', 'status_flag', 'expert','recruiter']
    search_fields = ['name', 'email', 'phone_number']
    autocomplete_fields = ['technology', 'expert','recruiter']

    @admin.display(description="Expert")
    def expert_status(self, obj):
        return f"✅ {obj.expert.name}" if obj.expert else "❌ No Expert"


# ---------- Feedback filter for Case ----------

class FeedbackStatusFilter(admin.SimpleListFilter):
    title = 'Feedback Status'
    parameter_name = 'feedback_status'

    def lookups(self, request, model_admin):
        return [
            ('done', 'Feedback done'),
            ('not_done', 'Not complete')
        ]

    def queryset(self, request, queryset):
        # Matches: feedback_status = 'Not complete' if len(strip(feedback)) < 5
        if self.value() == 'done':
            return queryset.filter(feedback__isnull=False).exclude(feedback__regex=r'^\s{0,4}$')
        if self.value() == 'not_done':
            return queryset.filter(feedback__isnull=True) | queryset.filter(feedback__regex=r'^\s{0,4}$')


# ---------- Case form with emoji labels for candidates ----------

class CandidateChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.name} {'✅' if obj.expert_id else '❌'}"

class CaseAdminForm(forms.ModelForm):
    class Meta:
        model = Case
        fields = "__all__"

    # Use a normal select so emojis appear in the dropdown list
    candidate = CandidateChoiceField(queryset=Candidate.objects.all())


# ---------- Case Admin ----------

@admin.register(Case)
class CaseAdmin(admin.ModelAdmin):
    form = CaseAdminForm

    list_display = [
        'id', 'company', 'candidate', 'support_type',
        'date', 'start_time', 'end_time', 'round',
        'expert', 'task', 'filled_by', 'status', 'feedback_status'
    ]
    list_filter = [
        'company', 'support_type', 'round',
        'date', 'expert', 'task', 'filled_by', 'candidate',
        FeedbackStatusFilter
    ]
    search_fields = [
        'company__company_name',  # FK text field
        'candidate__name',
        'expert__name',
        'feedback',
        'filled_by__username',
    ]
    # Keep autocomplete for other fields; NOT for candidate (to show emojis in dropdown)
    autocomplete_fields = ['expert', 'task', 'support_type', 'round', 'filled_by']
    readonly_fields = ['filled_by', 'candidate_technology', 'candidate_email', 'candidate_phone','candidate_resume','feedback','task','status']

    @admin.display(description='Feedback Status', ordering='feedback')
    def feedback_status(self, obj):
        text = (obj.feedback or '').strip()
        return 'Not complete' if len(text) < 5 else 'Feedback done'

    def save_model(self, request, obj, form, change):
        if not obj.filled_by:
            obj.filled_by = request.user
        super().save_model(request, obj, form, change)


    

@admin.register(PH)
class PHAdmin(admin.ModelAdmin):
    list_display = ['date', 'candidate', 'technology', 'company', 'expert', 'team_lead', 'po_type']
    readonly_fields = ['expert', 'team_lead']
    autocomplete_fields = ['candidate', 'technology']


# ---------- Notifications (read-only add) ----------

@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ('sender', 'recipient', 'message', 'created_at', 'read')
    list_filter = ('read', 'created_at')
    ordering = ('-created_at',)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Superusers only see notifications addressed to themselves
        if request.user.is_superuser:
            return qs.filter(recipient=request.user)
        return qs.none()

    def has_add_permission(self, request):
        return False  # Users cannot create notifications from admin


# ---------- Marketing Team ----------

class MarketingTeamAdminForm(forms.ModelForm):
    class Meta:
        model = MarketingTeam
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # ✅ Only marketing members who are marked as team leads can be selected
        self.fields['team_lead'].queryset = MarketingTeam.objects.filter(is_team_lead=True)


@admin.register(MarketingTeam)
class MarketingTeamAdmin(admin.ModelAdmin):
    form = MarketingTeamAdminForm   # ✅ use the custom form

    list_display = [
        'id', 'name', 'is_team_lead', 'team_lead',
        'date_of_joining', 'status_flag', 'team_members_list',
    ]
    list_filter = ['is_team_lead', 'status_flag', 'date_of_joining']
    search_fields = ['name', 'user__username', 'user__email']
    # 🚨 Important: remove team_lead from autocomplete if you want filtering to apply
    autocomplete_fields = ['user']

    def team_members_list(self, obj):
        return ", ".join(member.name for member in obj.team_members.all()) or "No members"
    team_members_list.short_description = "Team Members"



# ---------- Export all models to XLSX ----------

EXCLUDED_MODELS = ['session', 'contenttype', 'group', 'permission']

def export_all_models_xlsx(request):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    all_models = apps.get_models()

    def safe_value(val):
        if hasattr(val, '__str__'):
            return str(val)
        return val

    for model in all_models:
        model_name = model._meta.model_name.lower()
        if model_name in EXCLUDED_MODELS:
            continue

        queryset = model.objects.all()
        ws = wb.create_sheet(title=model_name[:31])

        if not queryset.exists():
            ws.append(['No data'])
            continue

        headers = [field.name for field in model._meta.fields]
        ws.append(headers)

        for obj in queryset:
            row = []
            for field in headers:
                val = getattr(obj, field)
                if hasattr(val, 'all'):
                    val = ", ".join(str(i) for i in val.all())
                else:
                    val = safe_value(val)
                row.append(val)
            ws.append(row)

        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=all_models_data.xlsx'
    wb.save(response)
    return response


def get_admin_urls(urls):
    def get_urls():
        custom_urls = [
            path('export-all-models-xlsx/', admin.site.admin_view(export_all_models_xlsx), name='export_all_models_xlsx'),
        ]
        return custom_urls + urls
    return get_urls

admin.site.get_urls = get_admin_urls(admin.site.get_urls())
